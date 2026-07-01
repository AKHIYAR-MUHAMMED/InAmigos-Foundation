import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_volunteer_sheet():
    # Create workbook and select active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Volunteer Initiatives"
    
    # Ensure grid lines are visible in the spreadsheet
    ws.views.sheetView[0].showGridLines = True
    
    # Custom colors & styling definitions
    NAVY_COLOR = "1A365D"       # #1A365D
    TEAL_ACCENT = "319795"      # #319795
    ZEBRA_COLOR = "F7FAFC"      # #F7FAFC (Light gray)
    BORDER_COLOR = "E2E8F0"     # #E2E8F0 (Slate border)
    
    # Border definition
    thin_side = Side(style='thin', color=BORDER_COLOR)
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # Headers
    headers = ["Organization Name", "Opportunity Type", "Eligibility & Requirements", "Application Link", "Deadline & Cycle"]
    
    # Write Header row
    ws.append(headers)
    
    # Style Header row
    ws.row_dimensions[1].height = 26
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=NAVY_COLOR, end_color=NAVY_COLOR, fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = cell_border
        
    # Opportunities dataset (17 total)
    opportunities = [
        {
            "org": "Teach For India",
            "type": "Fellowship (Full-time, 2 Years)",
            "eligibility": "Indian citizens or OCI holders. Must have completed graduation by June/July of the cohort year. Looking for strong leadership potential and commitment to educational equity.",
            "link": "https://www.teachforindia.org/",
            "deadline": "Cyclical annual rounds. Applications typically open in August/September and close in March (with rolling selection stages)."
        },
        {
            "org": "SBI Foundation",
            "type": "SBI Youth for India Fellowship (Full-time, 13 Months)",
            "eligibility": "Indian citizens, OCI, or citizens of Nepal/Bhutan aged 21-32 years at fellowship start. Must hold at least a Bachelor's degree. Passion for working on grassroots rural development.",
            "link": "https://www.youthforindia.org/",
            "deadline": "Cyclical annual batch. Applications typically open between March and May for the cohort starting in August."
        },
        {
            "org": "Kaivalya Education Foundation (KEF) / Piramal Foundation",
            "type": "Gandhi Fellowship (Full-time, 2 Years)",
            "eligibility": "Graduates or final-year students (under age 25, or up to 29 with 1 year work experience). Requires minimum 55% aggregate marks in 10th, 12th, and graduation.",
            "link": "https://gandhifellowship.org/",
            "deadline": "Cyclical annual cycle. Applications open around November and close by March 31st."
        },
        {
            "org": "PRS Legislative Research",
            "type": "LAMP Fellowship (Full-time, 11 Months)",
            "eligibility": "Indian citizens aged 25 years or below. Must hold at least a Bachelor’s degree in any discipline. Highly valuable for students interested in public policy, research, and law.",
            "link": "https://prsindia.org/lamp",
            "deadline": "Cyclical. Registration typically opens in October/November and closes by mid-December."
        },
        {
            "org": "NITI Aayog, Government of India",
            "type": "NITI Aayog Internship Scheme (Full-time, 1-3 Months)",
            "eligibility": "UG/PG students or research scholars enrolled in recognized institutions. Academic criteria: 85%+ in 12th for UG; 70%+ in graduation for PG; 70%+ in PG for PhD.",
            "link": "https://www.niti.gov.in/",
            "deadline": "Rolling monthly intake. The application portal opens exclusively from the 1st to the 10th of every month."
        },
        {
            "org": "Child Rights and You (CRY)",
            "type": "CRY Volunteering & Internships (Flexible)",
            "eligibility": "Open to students aged 14 and above (volunteering); college students/graduates (internships). Focuses on child rights advocacy, teaching, creative campaigns, and field support.",
            "link": "https://www.cry.org/",
            "deadline": "Rolling basis. Applications are accepted year-round depending on local chapter projects and campaigns."
        },
        {
            "org": "World Wide Fund for Nature (WWF) India",
            "type": "WWF Volunteer & Intern Program (Flexible)",
            "eligibility": "Volunteering: open to all citizens passionate about wildlife and environmental conservation. Internships: final-year UG or PG students with a formal recommendation letter.",
            "link": "https://www.wwfindia.org/",
            "deadline": "Volunteering is rolling. Internships open bi-annually (typically in February and October) or based on project requirements."
        },
        {
            "org": "Bhumi",
            "type": "Bhumi Fellowship & Volunteering (Flexible / Full-time)",
            "eligibility": "Fellowship: Graduates / young professionals aged 20-30. Volunteering: college students and professionals willing to support community service projects.",
            "link": "https://fellowship.bhumi.ngo/",
            "deadline": "Fellowship applications close in May annually. Weekend volunteering programs accept registrations year-round."
        },
        {
            "org": "Make A Difference (MAD)",
            "type": "MAD Youth Mentorship Volunteering (Flexible / 1-Year)",
            "eligibility": "College students or young professionals who can commit to 2-4 hours a week for a full year (52 weeks) to mentor children in shelter homes.",
            "link": "https://makeadiff.in/",
            "deadline": "Rolling. Recruitment campaigns are run locally in cities based on shelter requirements, usually active from May to August."
        },
        {
            "org": "eVidyaloka Trust",
            "type": "eVidyaloka Online Teaching Volunteer (Remote)",
            "eligibility": "Open to Indian citizens or OCI holders. Graduate in any discipline; must be fluent in a regional Indian language (Hindi, Tamil, Telugu, etc.); requires a laptop and good internet.",
            "link": "https://www.evidyaloka.org/",
            "deadline": "Rolling basis. Open all year; volunteers can teach 2 hours per week from home."
        },
        {
            "org": "Goonj",
            "type": "Goonj Volunteer & Intern Program (Flexible / 1 Month+)",
            "eligibility": "Open to school/college students and citizens. Internships require a full-time commitment of at least 1 month at Goonj centers (Mumbai, Delhi, Bangalore, etc.).",
            "link": "https://goonj.org/volunteer-intern/",
            "deadline": "Rolling. Open year-round. Applicants apply online or email 'mail@goonj.org' with a statement of purpose."
        },
        {
            "org": "Ministry of External Affairs, India",
            "type": "MEA Internship Scheme (Full-time, 1-3 Months)",
            "eligibility": "Indian citizens. Final-year undergraduate students or graduates from recognized universities, aged up to 25 years. Minimum 60% aggregate marks required.",
            "link": "https://internship.mea.gov.in/",
            "deadline": "Bi-annual. Two terms of intake. Term I applications usually close in March; Term II closes in August."
        },
        {
            "org": "Ashoka University",
            "type": "Young India Fellowship (YIF) (Full-time, 1 Year)",
            "eligibility": "Recognized undergraduate degree holders or final-year students in any discipline. Encourages diversity in academic background; no strict age limit.",
            "link": "https://apply.ashoka.edu.in",
            "deadline": "Cyclical. Applications open in October and close around late March in multiple rounds."
        },
        {
            "org": "Ministry of Education, Government of India",
            "type": "Prime Minister’s Research Fellowship (PMRF) (PhD Fellowship)",
            "eligibility": "PhD students in Science/Technology at PMRF-granting institutes (IITs, IISc, NITs, select Central Univs). Requires 8.0+ CGPA and nomination from the host institute.",
            "link": "https://may2024.pmrf.in/",
            "deadline": "Cyclical bi-annual cycle. Applications open through host institutes in May/June and November/December."
        },
        {
            "org": "National Foundation for India (NFI)",
            "type": "Abhijit Sen Rural Internship (ASRI) (Field Work, 45 Days)",
            "eligibility": "UG or PG students enrolled in any recognized Indian college or university. Focuses on research, reporting, and documenting rural economic issues.",
            "link": "https://www.nfi.org.in",
            "deadline": "Cyclical. Applications open around December/January and close by late February each year."
        },
        {
            "org": "The Robin Hood Army",
            "type": "Robin Hood Army Volunteer Program (Flexible)",
            "eligibility": "Zero-barrier entry. Open to anyone willing to devote time to collect surplus food from restaurants and distribute it to underprivileged communities.",
            "link": "https://robinhoodarmy.com",
            "deadline": "Rolling. Active year-round. Sign up anytime through the official website or WhatsApp links."
        },
        {
            "org": "United Nations Volunteers (UNV) India",
            "type": "UNV National/Community Volunteer (Flexible / Project-based)",
            "eligibility": "Minimum 18 years of age. Assignments range from digital marketing to field operations; specific qualifications depend on the active posting.",
            "link": "https://appwrite.unv.org",
            "deadline": "Rolling. General database registration is open year-round; specific project opportunities are posted with immediate deadlines."
        }
    ]
    
    # Styling definitions for data cells
    data_font = Font(name="Calibri", size=10, color="2D3748")
    link_font = Font(name="Calibri", size=10, color="0000FF", underline="single")
    
    left_align_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center_align_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    zebra_fill = PatternFill(start_color=ZEBRA_COLOR, end_color=ZEBRA_COLOR, fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Write opportunities to cells
    for idx, opt in enumerate(opportunities):
        row_num = idx + 2  # data starts at row 2
        
        ws.cell(row=row_num, column=1, value=opt["org"])
        ws.cell(row=row_num, column=2, value=opt["type"])
        ws.cell(row=row_num, column=3, value=opt["eligibility"])
        
        # Write Link as Excel Hyperlink
        link_cell = ws.cell(row=row_num, column=4, value=opt["link"])
        link_cell.hyperlink = opt["link"]
        
        ws.cell(row=row_num, column=5, value=opt["deadline"])
        
        # Determine background fill (Zebra striping)
        current_fill = zebra_fill if row_num % 2 == 0 else white_fill
        ws.row_dimensions[row_num].height = 42  # standard height to accommodate wrapped text nicely
        
        # Apply styles across columns
        for col_idx in range(1, 6):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.border = cell_border
            cell.fill = current_fill
            
            # Alignments
            if col_idx in [1, 2, 3]:
                cell.alignment = left_align_wrap
                cell.font = data_font
            elif col_idx == 4:
                cell.alignment = center_align_wrap
                cell.font = link_font
            elif col_idx == 5:
                cell.alignment = left_align_wrap
                cell.font = data_font
                
    # Define exact, visually optimal column widths (preventing truncation)
    col_widths = {
        "A": 26, # Org Name
        "B": 28, # Type
        "C": 55, # Eligibility
        "D": 28, # Link
        "E": 40  # Deadline/Cycle
    }
    
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width
        
    # Save Workbook
    filename = "Volunteer_Opportunities_India.xlsx"
    wb.save(filename)
    print(f"Excel Sheet generated successfully: {os.path.abspath(filename)}")

if __name__ == "__main__":
    create_volunteer_sheet()
